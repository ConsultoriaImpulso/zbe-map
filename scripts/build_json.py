import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
XLSX_FILE = BASE_DIR / "ZBE.xlsx"
MAP_FILE = BASE_DIR / "mapa_zbe.csv"
OUT_FILE = BASE_DIR / "access.json"


def log(msg: str):
    print(msg, flush=True)


def canon(text: str) -> str:
    text = str(text or "").strip().lower()
    repl = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ü": "u",
        "à": "a", "è": "e", "ì": "i", "ò": "o", "ù": "u",
    }
    for k, v in repl.items():
        text = text.replace(k, v)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def html_escape(text: str) -> str:
    return (
        str(text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def text_to_html(text: str) -> str:
    """
    Convierte texto plano a HTML simple:
    - preserva saltos de línea
    - preserva párrafos
    - convierte **texto** en <strong>texto</strong>
    """
    if text is None:
        return ""

    s = str(text)
    if not s.strip():
        return ""

    s = html_escape(s)

    # negritas estilo markdown
    s = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", s)

    s = s.replace("\r\n", "\n").replace("\r", "\n")

    blocks = [b.strip() for b in re.split(r"\n{2,}", s) if b.strip()]
    if not blocks:
        return ""

    html_blocks = []
    for block in blocks:
        html_blocks.append(f"<p>{block.replace(chr(10), '<br>')}</p>")

    return "".join(html_blocks)


def detect_columns(headers):
    normalized = [canon(h) for h in headers]

    def find(*names):
        for i, h in enumerate(normalized):
            for name in names:
                if canon(name) in h:
                    return i
        return -1

    cols = {
        "city": find("ciudad", "city", "municipio"),
        "badge": find("distintivo ambiental", "distintivo", "etiqueta"),
        "veh": find("tipo vehiculo", "tipo vehículo", "vehiculo", "vehículo", "vehicle"),
        "access": find("acceso", "access"),
        "obs": find("observaciones", "observacion", "observación", "obs", "nota"),
        "vig": find("estado zbe", "vigencia", "vigente", "estado"),
        "url": find("url", "fuente", "link", "enlace"),
    }
    return cols


def load_coords(csv_path: Path):
    coords = {}

    if not csv_path.exists():
        raise FileNotFoundError(f"No existe el archivo de coordenadas: {csv_path.name}")

    log(f"Leyendo coordenadas desde: {csv_path.name}")

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)

        delim = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delim)

        if not reader.fieldnames:
            raise ValueError("mapa_zbe.csv no tiene encabezados")

        headers = [canon(h) for h in reader.fieldnames]

        def find_field(*names):
            for h_raw in reader.fieldnames:
                h = canon(h_raw)
                for n in names:
                    if canon(n) in h:
                        return h_raw
            return None

        city_f = find_field("ciudad", "city", "municipio")
        lat_f = find_field("lat", "latitud", "latitude", "y")
        lng_f = find_field("lng", "longitud", "lon", "longitude", "x")

        log(f"Columnas mapa_zbe.csv: {reader.fieldnames}")
        log(f"Detectadas -> ciudad: {city_f}, lat: {lat_f}, lng: {lng_f}")

        if not city_f or not lat_f or not lng_f:
            raise ValueError("No encuentro en mapa_zbe.csv las columnas de ciudad/lat/lng")

        for row in reader:
            city = (row.get(city_f) or "").strip()
            lat = (row.get(lat_f) or "").strip()
            lng = (row.get(lng_f) or "").strip()

            if not city:
                continue

            try:
                lat_v = float(lat.replace(",", "."))
                lng_v = float(lng.replace(",", "."))
            except Exception:
                continue

            coords[canon(city)] = {
                "city": city,
                "lat": lat_v,
                "lng": lng_v
            }

    log(f"Coordenadas cargadas: {len(coords)}")
    return coords


def main():
    log("=== INICIO build_json ===")
    log(f"Base dir: {BASE_DIR}")
    log(f"Existe ZBE.xlsx: {XLSX_FILE.exists()}")
    log(f"Existe mapa_zbe.csv: {MAP_FILE.exists()}")

    if not XLSX_FILE.exists():
        raise FileNotFoundError("No existe ZBE.xlsx en la raíz del repositorio")

    coords = load_coords(MAP_FILE)

    log(f"Leyendo Excel: {XLSX_FILE.name}")
    wb = load_workbook(XLSX_FILE, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        raise ValueError("ZBE.xlsx está vacío")

    headers = [cell.value for cell in rows[0]]
    log(f"Encabezados Excel: {headers}")

    cols = detect_columns(headers)
    log(f"Columnas detectadas: {cols}")

    required = ["city", "badge", "veh", "access"]
    missing = [k for k in required if cols[k] < 0]
    if missing:
        raise ValueError(f"Faltan columnas obligatorias en ZBE.xlsx: {missing}")

    result = []
    rows_ok = 0

    for row in rows[1:]:
        city = row[cols["city"]].value if cols["city"] >= 0 else ""
        badge = row[cols["badge"]].value if cols["badge"] >= 0 else ""
        veh = row[cols["veh"]].value if cols["veh"] >= 0 else ""
        access = row[cols["access"]].value if cols["access"] >= 0 else ""

        if not city or not badge or not veh or not access:
            continue

        city_s = str(city).strip()
        badge_s = str(badge).strip()
        veh_s = str(veh).strip()
        access_s = str(access).strip()

        obs_html = ""
        if cols["obs"] >= 0:
            obs_val = row[cols["obs"]].value
            obs_html = text_to_html(obs_val)

        vig_s = ""
        if cols["vig"] >= 0 and row[cols["vig"]].value is not None:
            vig_s = str(row[cols["vig"]].value).strip()

        url_s = ""
        if cols["url"] >= 0 and row[cols["url"]].value is not None:
            url_s = str(row[cols["url"]].value).strip()

        coord = coords.get(canon(city_s), {})

        result.append({
            "city": city_s,
            "badge": badge_s,
            "vehicle": veh_s,
            "access": access_s,
            "observations_html": obs_html,
            "estado_zbe": vig_s,
            "url": url_s,
            "lat": coord.get("lat"),
            "lng": coord.get("lng")
        })
        rows_ok += 1

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    log(f"Filas exportadas: {rows_ok}")
    log(f"Archivo generado: {OUT_FILE.name}")
    log("=== FIN build_json OK ===")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log("=== ERROR build_json ===")
        log(f"Tipo: {type(e).__name__}")
        log(f"Mensaje: {e}")
        raise
