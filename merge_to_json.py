import csv
import json
import re
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
XLSX_FILE = BASE_DIR / "ZBE.xlsx"
MAP_FILE = BASE_DIR / "mapa_zbe.csv"
OUT_FILE = BASE_DIR / "access.json"


def canon(text: str) -> str:
    text = str(text or "").strip().lower()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ü": "u",
        "Á": "a", "É": "e", "Í": "i", "Ó": "o", "Ú": "u", "Ü": "u",
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    text = re.sub(r"\s+", " ", text)
    return text


def richtext_to_html(cell) -> str:
    """
    Convierte el contenido de una celda a HTML sencillo.
    Si hay rich text, intenta conservar negrita.
    Si no, conserva saltos de línea.
    """
    value = cell.value
    if value is None:
        return ""

    # Caso normal: texto plano
    if not hasattr(value, "runs"):
        text = str(value)
        text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        paragraphs = [p for p in text.split("\n\n")]
        if len(paragraphs) > 1:
            return "".join(f"<p>{p.replace(chr(10), '<br>')}</p>" for p in paragraphs)
        return text.replace("\n", "<br>")

    # Caso rich text
    parts = []
    for run in value.runs:
        txt = str(run.text or "")
        txt = txt.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        if run.font and run.font.b:
            txt = f"<strong>{txt}</strong>"
        parts.append(txt)

    html = "".join(parts)
    html = html.replace("\r\n", "\n").replace("\r", "\n")
    paragraphs = [p for p in html.split("\n\n")]
    if len(paragraphs) > 1:
        return "".join(f"<p>{p.replace(chr(10), '<br>')}</p>" for p in paragraphs)
    return html.replace("\n", "<br>")


def detect_columns(headers):
    h = [str(x).strip().lower() if x is not None else "" for x in headers]

    def find(*names):
        for i, col in enumerate(h):
            for name in names:
                if name in col:
                    return i
        return -1

    cols = {
        "city": find("ciudad", "city", "municipio"),
        "badge": find("distintivo", "etiqueta"),
        "veh": find("tipo vehículo", "tipo vehiculo", "vehículo", "vehiculo", "vehicle"),
        "access": find("acceso", "access"),
        "obs": find("observaciones", "observacion", "obs", "nota"),
        "vig": find("estado zbe", "vigencia", "vigente", "estado"),
        "url": find("url", "fuente", "link", "enlace"),
    }
    return cols


def load_coords(csv_path: Path):
    coords = {}
    if not csv_path.exists():
        return coords

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(2048)
        f.seek(0)

        delim = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delim)

        headers = [h.strip().lower() for h in (reader.fieldnames or [])]

        def find_field(*names):
            for h in headers:
                for n in names:
                    if n in h:
                        return h
            return None

        city_f = find_field("ciudad", "city", "municipio")
        lat_f = find_field("lat", "latitud", "latitude", "y")
        lng_f = find_field("lng", "long", "lon", "longitude", "x")

        for row in reader:
            city = row.get(city_f, "") if city_f else ""
            lat = row.get(lat_f, "") if lat_f else ""
            lng = row.get(lng_f, "") if lng_f else ""

            if not city:
                continue

            ccity = canon(city)
            try:
                lat_v = float(str(lat).replace(",", "."))
                lng_v = float(str(lng).replace(",", "."))
            except Exception:
                continue

            coords[ccity] = {
                "city": city.strip(),
                "lat": lat_v,
                "lng": lng_v
            }

    return coords


def main():
    if not XLSX_FILE.exists():
        raise FileNotFoundError(f"No existe {XLSX_FILE}")

    wb = load_workbook(XLSX_FILE, data_only=False, rich_text=True)
    ws = wb.active

    rows = list(ws.iter_rows())
    if not rows:
        raise ValueError("El Excel está vacío")

    headers = [cell.value for cell in rows[0]]
    cols = detect_columns(headers)

    required = ["city", "badge", "veh", "access"]
    missing = [k for k in required if cols[k] < 0]
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {missing}")

    coords = load_coords(MAP_FILE)

    result = []

    for row in rows[1:]:
        city = row[cols["city"]].value if cols["city"] >= 0 else ""
        badge = row[cols["badge"]].value if cols["badge"] >= 0 else ""
        veh = row[cols["veh"]].value if cols["veh"] >= 0 else ""
        access = row[cols["access"]].value if cols["access"] >= 0 else ""

        if not city or not badge or not veh or not access:
            continue

        obs_html = ""
        if cols["obs"] >= 0:
            obs_html = richtext_to_html(row[cols["obs"]])

        vig = ""
        if cols["vig"] >= 0 and row[cols["vig"]].value is not None:
            vig = str(row[cols["vig"]].value).strip()

        url = ""
        if cols["url"] >= 0 and row[cols["url"]].value is not None:
            url = str(row[cols["url"]].value).strip()

        ccity = canon(city)
        coord = coords.get(ccity, {})

        result.append({
            "city": str(city).strip(),
            "badge": str(badge).strip(),
            "vehicle": str(veh).strip(),
            "access": str(access).strip(),
            "observations_html": obs_html,
            "estado_zbe": vig,
            "url": url,
            "lat": coord.get("lat"),
            "lng": coord.get("lng")
        })

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"JSON generado correctamente: {OUT_FILE}")
    print(f"Filas exportadas: {len(result)}")


if __name__ == "__main__":
    main()
